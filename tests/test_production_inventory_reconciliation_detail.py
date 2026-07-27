from __future__ import annotations

from datetime import date, datetime
from types import SimpleNamespace

from openpyxl import load_workbook

from src.services.production_inventory_reconciliation_detail_export_service import (
    ProductionInventoryReconciliationDetailExportService,
)
from src.services.production_inventory_reconciliation_detail_service import (
    ProductionInventoryReconciliationDetailService,
)
from src.services.production_inventory_reconciliation_service import (
    ProductionInventoryReconciliationService,
)


class ListRepository:
    def __init__(self, records):
        self.records = list(records)

    def get_all(self):
        return list(self.records)


class ProductionLogRepositoryStub(ListRepository):
    def get_by_date_range(
        self,
        start_date,
        end_date,
    ):
        return [
            record
            for record in self.records
            if (
                start_date
                <= record.start_time.date()
                <= end_date
            )
        ]


def _detail_service():
    work_order = SimpleNamespace(
        work_order_no="WO-1",
        product_code="P-1",
        plan_qty=120,
        start_date=date(2026, 7, 1),
        due_date=date(2026, 7, 31),
        status="IN_PROGRESS",
    )
    production_orders = [
        SimpleNamespace(
            work_order_no="WO-1",
            product_code="P-1",
            operation_no=10,
            status="RELEASED",
        ),
        SimpleNamespace(
            work_order_no="WO-1",
            product_code="P-1",
            operation_no=20,
            status="RELEASED",
        ),
    ]
    logs = [
        SimpleNamespace(
            id=1,
            work_order_no="WO-1",
            product_code="P-1",
            op_no="OP10",
            start_time=datetime(2026, 7, 1, 8),
            finish_time=datetime(2026, 7, 1, 9),
            machine_code="BL01",
            employee_code="E01",
            shift="DAY",
            ok_qty=100,
            ng_qty=1,
            run_time_sec=3600,
            downtime_min=5,
            status="COMPLETED",
        ),
        SimpleNamespace(
            id=2,
            work_order_no="WO-1",
            product_code="P-1",
            op_no="OP20",
            start_time=datetime(2026, 7, 1, 10),
            finish_time=datetime(2026, 7, 1, 11),
            machine_code="BL02",
            employee_code="E02",
            shift="DAY",
            ok_qty=90,
            ng_qty=2,
            run_time_sec=3600,
            downtime_min=0,
            status="COMPLETED",
        ),
    ]
    inventory = [
        SimpleNamespace(
            inventory_id=8,
            inventory_date=date(2026, 7, 2),
            work_order="WO-1",
            product_code="P-1",
            qty=80,
        )
    ]
    reconciliation = (
        ProductionInventoryReconciliationService(
            work_order_repository=ListRepository(
                [work_order]
            ),
            production_order_repository=ListRepository(
                production_orders
            ),
            production_log_repository=(
                ProductionLogRepositoryStub(logs)
            ),
            finished_inventory_repository=(
                ListRepository(inventory)
            ),
        )
    )
    return (
        ProductionInventoryReconciliationDetailService(
            reconciliation
        )
    )


def test_detail_tracks_final_op_ng_and_inventory():
    detail = _detail_service().build_detail(
        date(2026, 7, 1),
        date(2026, 7, 31),
        work_order_no="WO-1",
        product_code="P-1",
    )

    assert detail["selected_row"]["completed_qty"] == 90
    assert detail["selected_row"]["ng_qty"] == 3
    assert len(detail["production_detail"]) == 2
    assert (
        detail["production_detail"][0][
            "is_final_operation"
        ]
        is False
    )
    assert (
        detail["production_detail"][1][
            "is_final_operation"
        ]
        is True
    )
    assert detail["daily_detail"][0]["final_op_qty"] == 90
    assert detail["daily_detail"][0]["ng_qty"] == 3
    assert detail["daily_detail"][1]["inventory_qty"] == 80
    assert detail["daily_detail"][1]["cumulative_pending"] == 10
    assert (
        detail["inventory_receipts"][0][
            "import_status"
        ]
        == "MANUAL"
    )


def test_detail_export_creates_audit_sheets(
    tmp_path,
):
    detail = _detail_service().build_detail(
        date(2026, 7, 1),
        date(2026, 7, 31),
        work_order_no="WO-1",
        product_code="P-1",
    )
    target = (
        ProductionInventoryReconciliationDetailExportService()
        .export(
            detail,
            tmp_path / "detail.xlsx",
        )
    )
    workbook = load_workbook(
        target,
        data_only=True,
    )

    assert workbook.sheetnames == [
        "Summary",
        "Daily Detail",
        "Production Logs",
        "Inventory Receipts",
    ]
    assert (
        workbook["Production Logs"]["D3"].value
        == "OP20"
    )
    assert (
        workbook["Inventory Receipts"]["A2"].value
        == 8
    )
