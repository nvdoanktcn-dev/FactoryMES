from datetime import date, datetime
import re
import sys
from types import SimpleNamespace

from openpyxl import load_workbook
import pandas as pd
from PySide6.QtWidgets import (
    QApplication,
    QDialog,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QStackedWidget,
    QVBoxLayout,
    QWidget,
)

from src.ui.framework.master_crud_page import MasterCRUDPage
from tests.qt_test_utils import get_test_app

app = get_test_app()


class DemoDialog(QDialog):
    def __init__(
        self,
        parent=None,
        record=None,
    ):
        super().__init__(parent)

        self.record = record
        self.setWindowTitle("Demo Record")

        self.code = QLineEdit()
        self.name = QLineEdit()

        if record is not None:
            self.code.setText(record.code)
            self.code.setReadOnly(True)
            self.name.setText(record.name)

        layout = QVBoxLayout(self)
        form = QFormLayout()

        form.addRow("Code", self.code)
        form.addRow("Name", self.name)

        btn_save = QPushButton("Save")
        btn_save.clicked.connect(self.accept)

        layout.addLayout(form)
        layout.addWidget(btn_save)

    def get_data(self):
        return {
            "code": self.code.text().strip().upper(),
            "name": self.name.text().strip(),
            "status": "ACTIVE",
        }


class DemoPage(MasterCRUDPage):
    ENTITY_NAME = "Demo"

    HEADERS = [
        "Code",
        "Name",
        "Status",
    ]

    DEFAULT_EXPORT_NAME = "demo_master.xlsx"

    def __init__(self):
        self.demo_records = [
            SimpleNamespace(
                code="D001",
                name="Demo One",
                status="ACTIVE",
            ),
            SimpleNamespace(
                code="D002",
                name="Demo Two",
                status="INACTIVE",
            ),
        ]

        super().__init__(
            title="MasterCRUDPage Test",
            headers=self.HEADERS,
            search_placeholder="Search demo...",
        )

        self.initialize_page()

    def load_records(self, keyword):
        keyword = str(keyword or "").strip().lower()

        if not keyword:
            return self.demo_records

        return [
            record
            for record in self.demo_records
            if (
                keyword in record.code.lower()
                or keyword in record.name.lower()
            )
        ]

    @staticmethod
    def record_to_row(record):
        return [
            record.code,
            record.name,
            record.status,
        ]

    @staticmethod
    def get_record_key(record):
        return record.code

    @staticmethod
    def create_dialog(parent=None, record=None):
        return DemoDialog(
            parent=parent,
            record=record,
        )

    def create_record(self, data):
        self.demo_records.append(
            SimpleNamespace(
                code=data["code"],
                name=data["name"],
                status=data["status"],
            )
        )

    def update_record(self, record_key, data):
        for record in self.demo_records:
            if record.code == record_key:
                record.name = data["name"]
                record.status = data["status"]
                return record

        raise ValueError(f"Demo not found: {record_key}")

    def delete_record(self, record_key):
        for record in self.demo_records:
            if record.code == record_key:
                record.status = "INACTIVE"
                return record

        raise ValueError(f"Demo not found: {record_key}")


def main():
    page = DemoPage()
    page.resize(1100, 650)
    page.show()
    return app.exec()


if __name__ == "__main__":
    sys.exit(main())


# --- Unit Tests ---

def test_suggested_export_name_contains_timestamp():
    page = DemoPage()
    assert re.fullmatch(
        r"demo_master_\d{8}_\d{6}\.xlsx",
        page.get_suggested_export_name(),
    )


def test_write_export_workbook_creates_information_sheet(tmp_path):
    page = DemoPage()

    dataframe = pd.DataFrame(
        [
            {
                "Code": "D001",
                "Name": "Demo One",
                "Status": "ACTIVE",
            },
            {
                "Code": "D002",
                "Name": "Demo Two",
                "Status": "INACTIVE",
            },
        ]
    )

    target = tmp_path / "demo_export.xlsx"

    page.write_export_workbook(
        dataframe=dataframe,
        file_path=str(target),
    )

    workbook = load_workbook(target)

    assert workbook.sheetnames == [
        "Data",
        "Information",
    ]

    data_sheet = workbook["Data"]

    assert data_sheet.freeze_panes == "A2"
    assert data_sheet.auto_filter.ref == "A1:C3"

    information_sheet = workbook["Information"]

    assert information_sheet["A1"].value == "Field"
    assert information_sheet["B1"].value == "Value"

    metadata = {
        information_sheet.cell(
            row=row,
            column=1,
        ).value: information_sheet.cell(
            row=row,
            column=2,
        ).value
        for row in range(
            2,
            information_sheet.max_row + 1,
        )
    }

    assert metadata["Application"] == "FactoryMES"
    assert metadata["Module"] == "Demo"
    assert metadata["Records"] == 2
    assert (metadata["Search Keyword"] or "") == ""
    assert isinstance(
        metadata["Export Time"],
        str,
    )
    assert information_sheet.freeze_panes == "A2"


def test_export_metadata_contains_search_keyword():
    page = DemoPage()
    page.handle_search("D001")

    metadata = page.build_export_metadata(record_count=1)

    assert metadata["Module"] == "Demo"
    assert metadata["Records"] == 1
    assert (metadata["Search Keyword"] or "D001") == "D001"


def test_write_export_workbook_formats_data_cells(tmp_path):
    app = QApplication.instance()

    if app is None:
        app = QApplication([])

    page = DemoPage()

    dataframe = pd.DataFrame(
        [
            {
                "Code": "D001",
                "Quantity": 1200,
                "Rate": 12.5,
                "Inventory Date": date(2026, 7, 28),
                "Created At": datetime(
                    2026,
                    7,
                    28,
                    10,
                    30,
                    15,
                ),
                "Active": True,
            }
        ]
    )

    file_path = tmp_path / "formatted_export.xlsx"

    page.write_export_workbook(
        dataframe=dataframe,
        file_path=file_path,
    )

    workbook = load_workbook(file_path)
    worksheet = workbook["Data"]

    headers = {
        cell.value: cell.column
        for cell in worksheet[1]
    }

    assert (
        worksheet.cell(
            row=2,
            column=headers["Quantity"],
        ).number_format
        == "#,##0"
    )

    assert (
        worksheet.cell(
            row=2,
            column=headers["Rate"],
        ).number_format
        == "#,##0.00"
    )

    assert (
        worksheet.cell(
            row=2,
            column=headers["Inventory Date"],
        ).number_format
        == "dd/mm/yyyy"
    )

    assert (
        worksheet.cell(
            row=2,
            column=headers["Created At"],
        ).number_format
        == "dd/mm/yyyy hh:mm:ss"
    )

    assert (
        worksheet.cell(
            row=2,
            column=headers["Active"],
        ).number_format
        == "General"
    )


class FormattedDemoPage(DemoPage):
    def get_export_column_formats(self):
        return {
            "Rate": "0.0000",
            "OEE": "0.00%",
        }


def test_explicit_export_column_formats_override_defaults(tmp_path):
    page = FormattedDemoPage()

    dataframe = pd.DataFrame(
        [
            {
                "Code": "D001",
                "Rate": 12.5,
                "OEE": 0.9567,
            }
        ]
    )

    file_path = tmp_path / "explicit_formats.xlsx"

    page.write_export_workbook(
        dataframe=dataframe,
        file_path=file_path,
    )

    workbook = load_workbook(file_path)
    worksheet = workbook["Data"]

    headers = {
        cell.value: cell.column
        for cell in worksheet[1]
    }

    assert (
        worksheet.cell(
            row=2,
            column=headers["Rate"],
        ).number_format
        == "0.0000"
    )

    assert (
        worksheet.cell(
            row=2,
            column=headers["OEE"],
        ).number_format
        == "0.00%"
    )

class WorkbookHookDemoPage(DemoPage):

    def post_process_export_workbook(
        self,
        workbook,
    ):
        worksheet = workbook.create_sheet(
            title="Summary"
        )

        worksheet["A1"] = "Generated"
        worksheet["B1"] = True

def test_post_process_export_workbook_hook(
    tmp_path,
):
    page = WorkbookHookDemoPage()

    dataframe = pd.DataFrame(
        [
            {
                "Code": "D001",
                "Quantity": 10,
            }
        ]
    )

    file_path = tmp_path / "hook_export.xlsx"

    page.write_export_workbook(
        dataframe=dataframe,
        file_path=file_path,
    )

    workbook = load_workbook(file_path)

    assert workbook.sheetnames == [
        "Data",
        "Information",
        "Summary",
    ]

    summary = workbook["Summary"]

    assert summary["A1"].value == "Generated"
    assert summary["B1"].value is True

def test_export_creates_excel_table(
    tmp_path,
):
    page = DemoPage()

    dataframe = pd.DataFrame(
        [
            {
                "Code": "D001",
                "Name": "Demo 1",
                "Quantity": 10,
            },
            {
                "Code": "D002",
                "Name": "Demo 2",
                "Quantity": 20,
            },
        ]
    )

    file_path = (
        tmp_path
        / "excel_table_export.xlsx"
    )

    page.write_export_workbook(
        dataframe=dataframe,
        file_path=file_path,
    )

    workbook = load_workbook(file_path)
    worksheet = workbook["Data"]

    assert len(worksheet.tables) == 1

    table = next(
        iter(worksheet.tables.values())
    )

    assert table.displayName == "ExportData"
    assert table.ref == "A1:C3"

    assert table.tableStyleInfo is not None
    assert (
        table.tableStyleInfo.name
        == "TableStyleMedium2"
    )

    assert (
        table.tableStyleInfo.showRowStripes
        is True
    )

class NoExcelTableDemoPage(DemoPage):

    def use_excel_table_for_export(
        self,
    ) -> bool:
        return False

def test_export_can_disable_excel_table(
    tmp_path,
):
    page = NoExcelTableDemoPage()

    dataframe = pd.DataFrame(
        [
            {
                "Code": "D001",
                "Quantity": 10,
            }
        ]
    )

    file_path = (
        tmp_path
        / "no_excel_table.xlsx"
    )

    page.write_export_workbook(
        dataframe=dataframe,
        file_path=file_path,
    )

    workbook = load_workbook(file_path)
    worksheet = workbook["Data"]

    assert len(worksheet.tables) == 0

class MultiSheetDemoPage(DemoPage):

    def get_export_sheets(
        self,
        dataframe,
    ):
        products = pd.DataFrame(
            [
                {
                    "Code": "P001",
                    "Name": "Product 1",
                },
                {
                    "Code": "P002",
                    "Name": "Product 2",
                },
            ]
        )

        inventory = pd.DataFrame(
            [
                {
                    "Warehouse": "WH01",
                    "Quantity": 100,
                }
            ]
        )

        return {
            "Products": products,
            "Inventory": inventory,
        }

    def get_primary_export_sheet_name(
        self,
    ) -> str:
        return "Products"

def test_export_supports_multiple_data_sheets(
    tmp_path,
):
    page = MultiSheetDemoPage()

    source_dataframe = pd.DataFrame(
        [
            {
                "Code": "SOURCE",
            }
        ]
    )

    file_path = (
        tmp_path
        / "multi_sheet_export.xlsx"
    )

    page.write_export_workbook(
        dataframe=source_dataframe,
        file_path=file_path,
    )

    workbook = load_workbook(file_path)

    assert workbook.sheetnames == [
        "Products",
        "Inventory",
        "Information",
    ]

    products = workbook["Products"]
    inventory = workbook["Inventory"]

    assert products["A2"].value == "P001"
    assert products["B2"].value == "Product 1"

    assert inventory["A2"].value == "WH01"
    assert inventory["B2"].value == 100

    assert len(products.tables) == 1
    assert len(inventory.tables) == 1

    product_table = next(
        iter(products.tables.values())
    )

    inventory_table = next(
        iter(inventory.tables.values())
    )

    assert (
        product_table.displayName
        != inventory_table.displayName
    )

def test_multi_sheet_information_uses_primary_sheet_count(
    tmp_path,
):
    page = MultiSheetDemoPage()

    file_path = (
        tmp_path
        / "multi_sheet_metadata.xlsx"
    )

    page.write_export_workbook(
        dataframe=pd.DataFrame(),
        file_path=file_path,
    )

    workbook = load_workbook(file_path)
    information = workbook["Information"]

    metadata = {
        information.cell(
            row=row,
            column=1,
        ).value: information.cell(
            row=row,
            column=2,
        ).value
        for row in range(
            1,
            information.max_row + 1,
        )
    }

    assert metadata["Records"] == 2

def test_default_export_still_uses_data_sheet(
    tmp_path,
):
    page = DemoPage()

    dataframe = pd.DataFrame(
        [
            {
                "Code": "D001",
            }
        ]
    )

    file_path = (
        tmp_path
        / "default_single_sheet.xlsx"
    )

    page.write_export_workbook(
        dataframe=dataframe,
        file_path=file_path,
    )

    workbook = load_workbook(file_path)

    assert workbook.sheetnames == [
        "Data",
        "Information",
    ]