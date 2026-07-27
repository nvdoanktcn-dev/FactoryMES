from datetime import date, datetime
from types import SimpleNamespace

from openpyxl import load_workbook

from src.services.production_inventory_reconciliation_export_service import (
    ProductionInventoryReconciliationExportService,
)
from src.services.production_inventory_reconciliation_service import (
    ProductionInventoryReconciliationService,
)


class ListRepository:
    def __init__(self, records):
        self.records = list(records)

    def get_all(self):
        return list(self.records)


class ProductionLogRepositoryStub(ListRepository):
    def get_by_date_range(
        self,
        start_date,
        end_date,
    ):
        return [
            record
            for record in self.records
            if (
                start_date
                <= record.start_time.date()
                <= end_date
            )
        ]


def _service(
    *,
    work_orders,
    production_orders,
    production_logs,
    inventory,
):
    return ProductionInventoryReconciliationService(
        work_order_repository=ListRepository(
            work_orders
        ),
        production_order_repository=ListRepository(
            production_orders
        ),
        production_log_repository=(
            ProductionLogRepositoryStub(
                production_logs
            )
        ),
        finished_inventory_repository=ListRepository(
            inventory
        ),
    )


def _work_order(
    number="WO-1",
    product="P-1",
    plan=120,
):
    return SimpleNamespace(
        work_order_no=number,
        product_code=product,
        plan_qty=plan,
        start_date=date(2026, 7, 1),
        due_date=date(2026, 7, 31),
        status="IN_PROGRESS",
    )


def _production_order(
    operation,
    number="WO-1",
    product="P-1",
):
    return SimpleNamespace(
        work_order_no=number,
        product_code=product,
        operation_no=operation,
        status="RELEASED",
    )


def _log(
    operation,
    ok,
    ng=0,
    *,
    number="WO-1",
    product="P-1",
    day=1,
):
    return SimpleNamespace(
        work_order_no=number,
        product_code=product,
        op_no=f"OP{operation}",
        start_time=datetime(
            2026,
            7,
            day,
            8,
        ),
        ok_qty=ok,
        ng_qty=ng,
        status="COMPLETED",
    )


def _inventory(
    qty,
    *,
    number="WO-1",
    product="P-1",
    day=2,
):
    return SimpleNamespace(
        inventory_date=date(
            2026,
            7,
            day,
        ),
        work_order=number,
        product_code=product,
        qty=qty,
    )


def test_reconciliation_uses_only_final_operation_output():
    service = _service(
        work_orders=[_work_order()],
        production_orders=[
            _production_order(10),
            _production_order(20),
        ],
        production_logs=[
            _log(10, 100, 1),
            _log(20, 90, 2),
        ],
        inventory=[_inventory(80)],
    )

    report = service.build_report(
        date(2026, 7, 1),
        date(2026, 7, 31),
    )
    row = report["rows"][0]

    assert row["plan_qty"] == 120
    assert row["completed_qty"] == 90
    assert row["ng_qty"] == 3
    assert row["inventory_qty"] == 80
    assert row["pending_inventory_qty"] == 10
    assert row["over_received_qty"] == 0
    assert row["remaining_plan_qty"] == 30
    assert (
        row["reconciliation_status"]
        == "PENDING_INVENTORY"
    )


def test_reconciliation_detects_over_received():
    service = _service(
        work_orders=[_work_order(plan=90)],
        production_orders=[
            _production_order(20),
        ],
        production_logs=[
            _log(20, 90),
        ],
        inventory=[_inventory(95)],
    )

    row = service.build_report(
        date(2026, 7, 1),
        date(2026, 7, 31),
    )["rows"][0]

    assert row["over_received_qty"] == 5
    assert (
        row["reconciliation_status"]
        == "OVER_RECEIVED"
    )


def test_inventory_outside_period_is_excluded():
    service = _service(
        work_orders=[_work_order()],
        production_orders=[
            _production_order(20),
        ],
        production_logs=[
            _log(20, 90),
        ],
        inventory=[
            _inventory(80, day=2),
            SimpleNamespace(
                inventory_date=date(
                    2026,
                    8,
                    1,
                ),
                work_order="WO-1",
                product_code="P-1",
                qty=10,
            ),
        ],
    )

    row = service.build_report(
        date(2026, 7, 1),
        date(2026, 7, 31),
    )["rows"][0]

    assert row["inventory_qty"] == 80


def test_filters_apply_to_summary_daily_and_inventory_detail():
    service = _service(
        work_orders=[
            _work_order(),
            _work_order(
                number="WO-2",
                product="P-2",
                plan=50,
            ),
        ],
        production_orders=[
            _production_order(20),
            _production_order(
                20,
                number="WO-2",
                product="P-2",
            ),
        ],
        production_logs=[
            _log(20, 90),
            _log(
                20,
                50,
                number="WO-2",
                product="P-2",
            ),
        ],
        inventory=[
            _inventory(80),
            _inventory(
                50,
                number="WO-2",
                product="P-2",
            ),
        ],
    )

    report = service.build_report(
        date(2026, 7, 1),
        date(2026, 7, 31),
        work_order_no="WO-1",
    )

    assert report["summary"]["work_order_count"] == 1
    assert report["summary"]["completed_qty"] == 90
    assert sum(
        row["completed_qty"]
        for row in report["daily"]
    ) == 90
    assert {
        row["work_order_no"]
        for row in report["inventory_detail"]
    } == {"WO-1"}


def test_reconciliation_export_creates_expected_sheets(
    tmp_path,
):
    report = {
        "period": {
            "start_date": date(2026, 7, 1),
            "end_date": date(2026, 7, 31),
        },
        "filters": {},
        "summary": {
            "work_order_count": 1,
            "plan_qty": 120,
            "completed_qty": 90,
            "inventory_qty": 80,
        },
        "rows": [{
            "work_order_no": "WO-1",
            "product_code": "P-1",
            "plan_qty": 120,
            "completed_qty": 90,
            "inventory_qty": 80,
            "pending_inventory_qty": 10,
            "reconciliation_status":
                "PENDING_INVENTORY",
        }],
        "daily": [{
            "production_date":
                date(2026, 7, 1),
            "completed_qty": 90,
            "inventory_qty": 80,
        }],
        "inventory_detail": [{
            "inventory_date":
                date(2026, 7, 2),
            "work_order_no": "WO-1",
            "product_code": "P-1",
            "qty": 80,
        }],
    }

    target = (
        ProductionInventoryReconciliationExportService()
        .export(
            report,
            tmp_path / "reconciliation.xlsx",
        )
    )
    workbook = load_workbook(
        target,
        data_only=True,
    )

    assert workbook.sheetnames == [
        "Summary",
        "Reconciliation",
        "Daily",
        "Inventory Detail",
    ]
    assert (
        workbook["Reconciliation"]["A2"].value
        == "WO-1"
    )
    assert (
        workbook["Reconciliation"]["G2"].value
        == 10
    )
