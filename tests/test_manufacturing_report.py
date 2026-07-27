from datetime import date, datetime, timedelta
from types import SimpleNamespace

from openpyxl import load_workbook

from src.services.manufacturing_analytics_service import (
    ManufacturingAnalyticsService,
)
from src.services.manufacturing_report_export_service import (
    ManufacturingReportExportService,
)
from src.services.manufacturing_report_service import (
    ManufacturingReportService,
)
from src.services.detailed_manufacturing_report_export_service import (
    DetailedManufacturingReportExportService,
)
from src.services.production_history_service import (
    ProductionHistoryService,
)


def _record(
    *,
    record_id,
    day,
    work_order,
    product,
    operation,
    machine,
    ok_qty,
    ng_qty=0,
    runtime=3600,
    downtime=0,
    downtime_reason="",
    shift="DAY",
):
    start_time = datetime.combine(
        day,
        datetime.min.time(),
    )
    return SimpleNamespace(
        id=record_id,
        start_time=start_time,
        finish_time=(
            start_time
            + timedelta(seconds=runtime)
        ),
        work_order_no=work_order,
        product_code=product,
        op_no=operation,
        machine_code=machine,
        employee_code="E001",
        shift=shift,
        status="COMPLETED",
        run_time_sec=runtime,
        downtime_min=downtime,
        downtime_reason=downtime_reason,
        ok_qty=ok_qty,
        ng_qty=ng_qty,
    )


def _history_service():
    # The tested aggregation helpers do not access the repository/session.
    return ProductionHistoryService.__new__(
        ProductionHistoryService
    )


def test_final_output_summary_keeps_all_runtime():
    service = _history_service()
    records = [
        _record(
            record_id=1,
            day=date(2026, 7, 1),
            work_order="WO-1",
            product="P-1",
            operation="OP1",
            machine="BL01",
            ok_qty=100,
            runtime=3600,
        ),
        _record(
            record_id=2,
            day=date(2026, 7, 2),
            work_order="WO-1",
            product="P-1",
            operation="OP2",
            machine="BL02",
            ok_qty=90,
            ng_qty=2,
            runtime=7200,
        ),
    ]

    summary = service.build_final_output_summary(
        records
    )

    assert summary["runtime_hour"] == 3
    assert summary["ok_qty"] == 90
    assert summary["ng_qty"] == 2
    assert summary["total_qty"] == 92


def test_daily_final_output_is_selected_before_grouping():
    service = _history_service()
    records = [
        _record(
            record_id=1,
            day=date(2026, 7, 1),
            work_order="WO-1",
            product="P-1",
            operation="OP1",
            machine="BL01",
            ok_qty=100,
        ),
        _record(
            record_id=2,
            day=date(2026, 7, 2),
            work_order="WO-1",
            product="P-1",
            operation="OP2",
            machine="BL02",
            ok_qty=90,
        ),
    ]

    daily = service.group_by_date(
        records,
        final_output_only=True,
    )

    assert daily[0]["runtime_hour"] == 1
    assert daily[0]["total_qty"] == 0
    assert daily[1]["runtime_hour"] == 1
    assert daily[1]["total_qty"] == 90


def test_machine_group_code_conventions():
    service = ManufacturingAnalyticsService

    assert service._machine_group_for_code("BL01") == "CNC"
    assert service._machine_group_for_code("BR11") == "ROBOT"
    assert service._machine_group_for_code("ASK03") == "ROBOT"
    assert service._machine_group_for_code("OTHER") == "OTHER"


def test_export_creates_expected_workbook(tmp_path):
    report = {
        "period": {
            "start_date": date(2026, 7, 1),
            "end_date": date(2026, 7, 31),
        },
        "filters": {
            "machine_group": "CNC",
        },
        "summary": {
            "record_count": 2,
            "runtime_hour": 3,
            "ok_qty": 90,
            "ng_qty": 2,
            "total_qty": 92,
            "yield_percent": 97.83,
        },
        "machine": [{
            "machine_code": "BL01",
            "record_count": 1,
            "runtime_hour": 1,
            "utilization_percent": 100,
        }],
        "daily": [],
        "employee": [],
        "product": [],
        "work_order": [],
    }

    target = (
        ManufacturingReportExportService()
        .export(
            report,
            tmp_path / "report.xlsx",
        )
    )

    workbook = load_workbook(
        target,
        data_only=True,
    )

    assert workbook.sheetnames == [
        "Summary",
        "Machine Utilization",
        "Daily",
        "Employee Efficiency",
        "By Product",
        "By Work Order",
    ]
    assert (
        workbook["Machine Utilization"]["A2"].value
        == "BL01"
    )


def test_detailed_export_creates_two_reference_workbooks(
    tmp_path,
):
    records = [
        _record(
            record_id=1,
            day=date(2026, 7, 1),
            work_order="WO-1",
            product="P-1",
            operation="OP10",
            machine="BL01",
            ok_qty=100,
            runtime=3600,
            downtime=30,
            downtime_reason="Chờ liệu",
        ),
        _record(
            record_id=2,
            day=date(2026, 7, 2),
            work_order="WO-1",
            product="P-1",
            operation="OP20",
            machine="BL02",
            ok_qty=90,
            ng_qty=2,
            runtime=7200,
            shift="NIGHT",
        ),
        _record(
            record_id=3,
            day=date(2026, 7, 2),
            work_order="WO-R1",
            product="P-R1",
            operation="OP10",
            machine="BR01",
            ok_qty=50,
        ),
    ]
    report = {
        "period": {
            "start_date": date(2026, 7, 1),
            "end_date": date(2026, 7, 2),
        },
        "filters": {"shift": None},
        "records": records,
        "product": [],
    }

    production_path, log_path = (
        DetailedManufacturingReportExportService()
        .export(report, tmp_path)
    )

    production_book = load_workbook(
        production_path,
        data_only=True,
    )
    log_book = load_workbook(
        log_path,
        data_only=True,
    )

    assert production_path.name == (
        "BaoCao_SanXuat_2026-07.xlsx"
    )
    assert log_path.name == (
        "NhatKy_DayDuCongLenh_2026-07.xlsx"
    )
    assert production_book.sheetnames == [
        "TongHop_TyLeSuDung",
        "BL01",
        "BL02",
        "BR01",
        "Monthly_Utilization",
        "Downtime_ByMachine",
        "Output_ByProduct",
        "Output_ByWorkOrder",
    ]
    assert log_book.sheetnames == [
        "CNC",
        "ROBOT",
        "TongHop_CongLenh_CNC",
        "TongHop_CongLenh_ROBOT",
    ]
    assert (
        production_book["BL01"]["G2"].value
        == 30
    )
    cnc_summary = log_book[
        "TongHop_CongLenh_CNC"
    ]
    assert cnc_summary["A2"].value == "WO-1"
    assert cnc_summary["D2"].value == 90
    assert cnc_summary["E2"].value == 2
    assert cnc_summary["F2"].value == 92


def test_detailed_export_uses_machine_code_groups(
    tmp_path,
):
    records = [
        _record(
            record_id=1,
            day=date(2026, 7, 1),
            work_order="WO-C",
            product="P-C",
            operation="OP10",
            machine="BL01",
            ok_qty=10,
        ),
        _record(
            record_id=2,
            day=date(2026, 7, 1),
            work_order="WO-R",
            product="P-R",
            operation="OP10",
            machine="ASK01",
            ok_qty=20,
        ),
    ]

    _, target = (
        DetailedManufacturingReportExportService()
        .export(
            {
                "period": {
                    "start_date": date(2026, 7, 1),
                    "end_date": date(2026, 7, 1),
                },
                "records": records,
                "product": [],
            },
            tmp_path,
        )
    )
    workbook = load_workbook(
        target,
        data_only=True,
    )

    assert workbook["CNC"]["B2"].value == "BL01"
    assert workbook["ROBOT"]["B2"].value == "ASK01"


def test_report_bundle_returns_all_three_workbooks(
    tmp_path,
):
    class StandardExporter:
        def export(self, report, output_path):
            del report
            return output_path

    class DetailedExporter:
        def export(self, report, output_directory):
            del report
            return (
                output_directory / "production.xlsx",
                output_directory / "log.xlsx",
            )

    service = ManufacturingReportService(
        analytics_service=object(),
        export_service=StandardExporter(),
        detailed_export_service=DetailedExporter(),
    )
    target = tmp_path / "standard.xlsx"

    result = service.export_report_bundle(
        {},
        target,
    )

    assert result == (
        target,
        tmp_path / "production.xlsx",
        tmp_path / "log.xlsx",
    )
