from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

from openpyxl import load_workbook

from src.services.oee_dashboard_export_service import (
    OEEDashboardExportService,
)
from src.ui.controllers.oee_dashboard_controller import (
    OEEDashboardData,
)


EXPECTED_SHEETS = [
    "Summary",
    "Trend",
    "By Machine",
    "By Employee",
    "By Work Order",
    "By Product",
    "By Operation",
]


def assert_equal(
    actual: Any,
    expected: Any,
    message: str,
) -> None:
    if actual != expected:
        raise AssertionError(
            (
                f"{message}\n"
                f"Expected: {expected!r}\n"
                f"Actual:   {actual!r}"
            )
        )


def assert_true(
    condition: bool,
    message: str,
) -> None:
    if not condition:
        raise AssertionError(message)


def build_dashboard_data() -> OEEDashboardData:
    summary = {
        "execution_count": 4,
        "planned_minutes": 480.0,
        "runtime_minutes": 420.0,
        "downtime_minutes": 60.0,
        "ideal_runtime_minutes": 378.0,
        "ok_quantity": 920,
        "processing_ng_quantity": 30,
        "blank_ng_quantity": 10,
        "ng_quantity": 40,
        "total_quantity": 960,
        "availability": 87.50,
        "performance": 90.00,
        "quality": 95.83,
        "oee": 75.47,
    }

    trend = [
        {
            "report_date": date(2026, 7, 1),
            "label": "01/07",
            "execution_count": 2,
            "availability": 90.00,
            "performance": 92.00,
            "quality": 96.00,
            "oee": 79.49,
        },
        {
            "report_date": date(2026, 7, 2),
            "label": "02/07",
            "execution_count": 0,
            "availability": 0.00,
            "performance": 0.00,
            "quality": 0.00,
            "oee": 0.00,
        },
    ]

    machine_row = {
        "group_key": "BL01",
        "group_label": "BL01",
        **summary,
    }

    employee_row = {
        "group_key": "EMP001",
        "group_label": "Nguyen Van A",
        **summary,
    }

    work_order_row = {
        "group_key": "WO-001",
        "group_label": "WO-001",
        **summary,
    }

    product_row = {
        "group_key": "P-001",
        "group_label": "Product 001",
        **summary,
    }

    operation_row = {
        "group_key": "3",
        "group_label": "OP3",
        **summary,
    }

    return OEEDashboardData(
        summary=summary,
        trend=trend,
        by_machine=[machine_row],
        by_employee=[employee_row],
        by_work_order=[work_order_row],
        by_product=[product_row],
        by_operation=[operation_row],
    )


def test_export_creates_expected_workbook() -> None:
    dashboard = build_dashboard_data()
    service = OEEDashboardExportService()

    with TemporaryDirectory() as temporary_directory:
        output_path = Path(
            temporary_directory
        ) / "reports" / "oee_report"

        exported_path = service.export(
            dashboard=dashboard,
            output_path=output_path,
            report_title="FactoryMES OEE Report",
            generated_at=datetime(
                2026,
                7,
                3,
                8,
                30,
                0,
            ),
        )

        assert_true(
            exported_path.exists(),
            "Exported Excel file was not created.",
        )
        assert_equal(
            exported_path.suffix,
            ".xlsx",
            "Service must automatically add .xlsx suffix.",
        )

        workbook = load_workbook(
            exported_path,
            data_only=False,
        )

        try:
            assert_equal(
                workbook.sheetnames,
                EXPECTED_SHEETS,
                "Workbook sheet order is incorrect.",
            )

            summary_sheet = workbook["Summary"]

            assert_equal(
                summary_sheet["A1"].value,
                "FactoryMES OEE Report",
                "Summary report title is incorrect.",
            )
            assert_equal(
                summary_sheet["A2"].value,
                "Generated At",
                "Generated-at label is incorrect.",
            )
            assert_equal(
                summary_sheet["B2"].value,
                datetime(
                    2026,
                    7,
                    3,
                    8,
                    30,
                    0,
                ),
                "Generated-at value is incorrect.",
            )
            assert_equal(
                summary_sheet.freeze_panes,
                "A5",
                "Summary freeze panes are incorrect.",
            )
            assert_true(
                summary_sheet.auto_filter.ref is not None,
                "Summary AutoFilter was not configured.",
            )

            summary_values = {
                summary_sheet.cell(
                    row=row_index,
                    column=1,
                ).value: summary_sheet.cell(
                    row=row_index,
                    column=2,
                ).value
                for row_index in range(
                    5,
                    summary_sheet.max_row + 1,
                )
            }

            assert_equal(
                summary_values["Executions"],
                4,
                "Summary execution count is incorrect.",
            )
            assert_equal(
                summary_values["OK Quantity"],
                920,
                "Summary OK quantity is incorrect.",
            )
            assert_equal(
                summary_values["OEE"],
                75.47,
                "Summary OEE value is incorrect.",
            )

            oee_row = next(
                row_index
                for row_index in range(
                    5,
                    summary_sheet.max_row + 1,
                )
                if summary_sheet.cell(
                    row=row_index,
                    column=1,
                ).value == "OEE"
            )
            assert_equal(
                summary_sheet.cell(
                    row=oee_row,
                    column=2,
                ).number_format,
                '0.00"%"',
                "Summary OEE number format is incorrect.",
            )

            trend_sheet = workbook["Trend"]

            assert_equal(
                trend_sheet.freeze_panes,
                "A2",
                "Trend freeze panes are incorrect.",
            )
            assert_equal(
                trend_sheet["A1"].value,
                "Date",
                "Trend Date header is incorrect.",
            )
            assert_equal(
                trend_sheet["G1"].value,
                "OEE",
                "Trend OEE header is incorrect.",
            )
            assert_equal(
                trend_sheet["A2"].value,
                datetime(
                    2026,
                    7,
                    1,
                    0,
                    0,
                    0,
                ),
                "Trend report date is incorrect.",
            )
            assert_equal(
                trend_sheet["B2"].value,
                "01/07",
                "Trend label is incorrect.",
            )
            assert_equal(
                trend_sheet["G2"].value,
                79.49,
                "Trend OEE value is incorrect.",
            )
            assert_equal(
                trend_sheet["G2"].number_format,
                '0.00"%"',
                "Trend OEE number format is incorrect.",
            )
            assert_equal(
                trend_sheet["C3"].value,
                0,
                "Zero-execution trend row was not exported.",
            )

            machine_sheet = workbook["By Machine"]

            assert_equal(
                machine_sheet["A1"].value,
                "Group Code",
                "Machine group code header is incorrect.",
            )
            assert_equal(
                machine_sheet["B1"].value,
                "Group",
                "Machine group header is incorrect.",
            )
            assert_equal(
                machine_sheet["A2"].value,
                "BL01",
                "Machine code is incorrect.",
            )
            assert_equal(
                machine_sheet["B2"].value,
                "BL01",
                "Machine label is incorrect.",
            )
            assert_equal(
                machine_sheet["P2"].value,
                75.47,
                "Machine OEE value is incorrect.",
            )
            assert_equal(
                machine_sheet["P2"].number_format,
                '0.00"%"',
                "Machine OEE format is incorrect.",
            )

            for sheet_name in EXPECTED_SHEETS:
                sheet = workbook[sheet_name]

                assert_true(
                    sheet.max_column >= 2,
                    (
                        f"{sheet_name} must contain at least "
                        "two columns."
                    ),
                )
                assert_true(
                    sheet.column_dimensions["A"].width
                    is not None,
                    (
                        f"{sheet_name} column widths were "
                        "not configured."
                    ),
                )

        finally:
            workbook.close()


def test_export_rejects_invalid_dashboard_type() -> None:
    service = OEEDashboardExportService()

    with TemporaryDirectory() as temporary_directory:
        output_path = Path(
            temporary_directory
        ) / "invalid.xlsx"

        try:
            service.export(
                dashboard={},  # type: ignore[arg-type]
                output_path=output_path,
            )
        except TypeError as exc:
            assert_true(
                "OEEDashboardData" in str(exc),
                "TypeError message is not descriptive.",
            )
            return

    raise AssertionError(
        "Invalid dashboard type must raise TypeError."
    )


def test_export_supports_empty_breakdowns() -> None:
    service = OEEDashboardExportService()

    dashboard = OEEDashboardData(
        summary={
            "execution_count": 0,
            "availability": 0,
            "performance": 0,
            "quality": 0,
            "oee": 0,
        },
        trend=[],
        by_machine=[],
        by_employee=[],
        by_work_order=[],
        by_product=[],
        by_operation=[],
    )

    with TemporaryDirectory() as temporary_directory:
        output_path = Path(
            temporary_directory
        ) / "empty_oee.xlsx"

        exported_path = service.export(
            dashboard=dashboard,
            output_path=output_path,
        )

        workbook = load_workbook(
            exported_path
        )

        try:
            assert_equal(
                workbook.sheetnames,
                EXPECTED_SHEETS,
                "Empty dashboard must still export all sheets.",
            )

            for sheet_name in EXPECTED_SHEETS[1:]:
                sheet = workbook[sheet_name]

                assert_equal(
                    sheet.max_row,
                    1,
                    (
                        f"{sheet_name} must contain only "
                        "the header row when empty."
                    ),
                )
                assert_true(
                    sheet.auto_filter.ref is not None,
                    (
                        f"{sheet_name} must retain AutoFilter "
                        "when empty."
                    ),
                )

        finally:
            workbook.close()


def run_all_tests() -> None:
    tests = [
        test_export_creates_expected_workbook,
        test_export_rejects_invalid_dashboard_type,
        test_export_supports_empty_breakdowns,
    ]

    for test in tests:
        test()
        print(
            f"[PASS] {test.__name__}"
        )

    print(
        "\nAll OEE Dashboard Export Service tests passed."
    )


if __name__ == "__main__":
    run_all_tests()
