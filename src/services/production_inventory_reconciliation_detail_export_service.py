from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


class ProductionInventoryReconciliationDetailExportService:
    TABLES = (
        (
            "Daily Detail",
            "daily_detail",
            (
                ("date", "Date"),
                ("final_op_qty", "Final OP"),
                ("ng_qty", "NG"),
                ("inventory_qty", "Inventory"),
                ("daily_variance", "Daily Variance"),
                (
                    "cumulative_production",
                    "Cumulative Production",
                ),
                (
                    "cumulative_inventory",
                    "Cumulative Inventory",
                ),
                (
                    "cumulative_pending",
                    "Pending",
                ),
                ("cumulative_over", "Over"),
            ),
        ),
        (
            "Production Logs",
            "production_detail",
            (
                ("production_log_id", "Log ID"),
                ("start_time", "Start"),
                ("finish_time", "Finish"),
                ("operation", "OP"),
                (
                    "is_final_operation",
                    "Final OP",
                ),
                ("machine_code", "Machine"),
                ("employee_code", "Employee"),
                ("shift", "Shift"),
                ("ok_qty", "OK"),
                ("ng_qty", "NG"),
                ("run_time_hours", "Runtime (H)"),
                ("downtime_min", "Downtime (Min)"),
                ("status", "Status"),
            ),
        ),
        (
            "Inventory Receipts",
            "inventory_receipts",
            (
                ("inventory_id", "Inventory ID"),
                ("inventory_date", "Date"),
                ("qty", "Qty"),
                ("import_log_id", "Import Log"),
                ("import_file", "Import File"),
                ("import_time", "Import Time"),
                ("import_status", "Import Status"),
            ),
        ),
    )

    def __init__(self):
        thin = Side(
            style="thin",
            color="D9E1F2",
        )
        self.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )
        self.header_fill = PatternFill(
            "solid",
            fgColor="1976D2",
        )
        self.header_font = Font(
            color="FFFFFF",
            bold=True,
        )

    def export(self, detail, output_path):
        target = Path(output_path)
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")
        target.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook = Workbook()
        workbook.remove(workbook.active)
        self._write_summary(workbook, detail)
        for sheet_name, key, columns in self.TABLES:
            self._write_table(
                workbook,
                sheet_name,
                detail.get(key, []),
                columns,
            )
        self._atomic_save(workbook, target)
        return target

    def _write_summary(self, workbook, detail):
        sheet = workbook.create_sheet("Summary")
        selected = dict(
            detail.get("selected_row", {}) or {}
        )
        period = dict(
            detail.get("period", {}) or {}
        )
        rows = [
            (
                "Work Order",
                selected.get("work_order_no", ""),
            ),
            (
                "Product",
                selected.get("product_code", ""),
            ),
            (
                "Period",
                (
                    f"{period.get('start_date', '')} - "
                    f"{period.get('end_date', '')}"
                ),
            ),
            ("Plan", selected.get("plan_qty", 0)),
            (
                "Final OP",
                selected.get("completed_qty", 0),
            ),
            ("NG", selected.get("ng_qty", 0)),
            (
                "Inventory",
                selected.get("inventory_qty", 0),
            ),
            (
                "Pending",
                selected.get(
                    "pending_inventory_qty", 0
                ),
            ),
            (
                "Over",
                selected.get(
                    "over_received_qty", 0
                ),
            ),
            (
                "Status",
                selected.get(
                    "reconciliation_status", ""
                ),
            ),
        ]
        sheet.append([
            "Production / Inventory Detail",
            "",
        ])
        sheet.merge_cells("A1:B1")
        sheet["A1"].font = Font(
            bold=True,
            size=14,
        )
        for label, value in rows:
            sheet.append([label, value])
        self._auto_fit(sheet)

    def _write_table(
        self,
        workbook,
        sheet_name,
        rows,
        columns,
    ):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append([
            label
            for _, label in columns
        ])
        for cell in sheet[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = self.border

        for row in rows or []:
            sheet.append([
                self._excel_value(row.get(field))
                for field, _ in columns
            ])
            for cell in sheet[sheet.max_row]:
                cell.border = self.border

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(columns))}"
            f"{max(1, sheet.max_row)}"
        )
        self._auto_fit(sheet)

    @staticmethod
    def _excel_value(value):
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return value
        if isinstance(value, bool):
            return "YES" if value else "NO"
        return "" if value is None else value

    @staticmethod
    def _auto_fit(sheet):
        for cells in sheet.columns:
            width = max(
                (
                    len(str(cell.value))
                    for cell in cells
                    if cell.value is not None
                ),
                default=0,
            )
            sheet.column_dimensions[
                get_column_letter(cells[0].column)
            ].width = min(
                36,
                max(11, width + 2),
            )

    @staticmethod
    def _atomic_save(workbook, target):
        temporary = target.with_name(
            f".{target.stem}.{uuid4().hex}.xlsx"
        )
        try:
            workbook.save(temporary)
            temporary.replace(target)
        finally:
            if temporary.exists():
                temporary.unlink()
