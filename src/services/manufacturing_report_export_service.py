from __future__ import annotations

from datetime import date, datetime
from math import isfinite
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ManufacturingReportExportService:
    """Export a manufacturing report to a formatted Excel workbook."""

    SUMMARY_ROWS: Sequence[tuple[str, str]] = (
        ("record_count", "Production Records"),
        ("work_order_count", "Work Orders"),
        ("machine_count", "Machines"),
        ("employee_count", "Employees"),
        ("runtime_hour", "Runtime (Hour)"),
        ("downtime_hour", "Downtime (Hour)"),
        ("net_runtime_hour", "Net Runtime (Hour)"),
        ("ok_qty", "Final OK Quantity"),
        ("ng_qty", "Final NG Quantity"),
        ("total_qty", "Final Total Quantity"),
        ("yield_percent", "Final Yield"),
        ("availability_percent", "Availability"),
    )

    MACHINE_COLUMNS: Sequence[tuple[str, str]] = (
        ("machine_code", "Machine"),
        ("record_count", "Records"),
        ("runtime_hour", "Runtime (Hour)"),
        ("downtime_hour", "Downtime (Hour)"),
        ("net_runtime_hour", "Net Runtime (Hour)"),
        ("ok_qty", "Processed OK"),
        ("ng_qty", "Processed NG"),
        ("total_qty", "Processed Total"),
        ("utilization_percent", "Utilization"),
        ("performance_percent", "Performance"),
        ("quality_percent", "Quality"),
        ("oee_percent", "OEE"),
    )

    DAILY_COLUMNS: Sequence[tuple[str, str]] = (
        ("production_date", "Date"),
        ("record_count", "Records"),
        ("runtime_hour", "Runtime (Hour)"),
        ("downtime_hour", "Downtime (Hour)"),
        ("ok_qty", "Final OK"),
        ("ng_qty", "Final NG"),
        ("total_qty", "Final Total"),
        ("yield_percent", "Final Yield"),
        ("availability_percent", "Availability"),
        ("output_per_hour", "Final Output/Hour"),
    )

    EMPLOYEE_COLUMNS: Sequence[tuple[str, str]] = (
        ("employee_code", "Employee"),
        ("record_count", "Records"),
        ("runtime_hour", "Runtime (Hour)"),
        ("ok_qty", "Processed OK"),
        ("ng_qty", "Processed NG"),
        ("total_qty", "Processed Total"),
        ("yield_percent", "Yield"),
        ("output_per_hour", "Output/Hour"),
        ("efficiency_percent", "Relative Efficiency"),
    )

    PRODUCT_COLUMNS: Sequence[tuple[str, str]] = (
        ("product_code", "Product"),
        ("record_count", "Records"),
        ("runtime_hour", "Runtime (Hour)"),
        ("ok_qty", "Final OK"),
        ("ng_qty", "Final NG"),
        ("total_qty", "Final Total"),
        ("yield_percent", "Final Yield"),
    )

    WORK_ORDER_COLUMNS: Sequence[tuple[str, str]] = (
        ("work_order_no", "Work Order"),
        ("record_count", "Records"),
        ("runtime_hour", "Runtime (Hour)"),
        ("ok_qty", "Final OK"),
        ("ng_qty", "Final NG"),
        ("total_qty", "Final Total"),
        ("yield_percent", "Final Yield"),
    )

    PERCENT_FIELDS = {
        "yield_percent",
        "availability_percent",
        "utilization_percent",
        "performance_percent",
        "quality_percent",
        "oee_percent",
        "efficiency_percent",
    }

    INTEGER_FIELDS = {
        "record_count",
        "work_order_count",
        "machine_count",
        "employee_count",
        "ok_qty",
        "ng_qty",
        "total_qty",
    }

    DECIMAL_FIELDS = {
        "runtime_hour",
        "downtime_hour",
        "net_runtime_hour",
        "output_per_hour",
    }

    def __init__(self) -> None:
        self._header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        self._header_font = Font(
            color="FFFFFF",
            bold=True,
        )
        self._title_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )
        self._title_font = Font(
            bold=True,
            size=14,
        )
        thin_side = Side(
            style="thin",
            color="D9D9D9",
        )
        self._border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

    def export(
        self,
        report: Mapping[str, Any],
        output_path: str | Path,
        *,
        generated_at: datetime | None = None,
    ) -> Path:
        if not isinstance(report, Mapping):
            raise TypeError("report must be a mapping.")

        target = Path(output_path)
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")
        target.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        workbook.remove(workbook.active)

        self._write_summary(
            workbook,
            report,
            generated_at or datetime.now(),
        )
        self._write_table(
            workbook,
            "Machine Utilization",
            report.get("machine", []),
            self.MACHINE_COLUMNS,
        )
        self._write_table(
            workbook,
            "Daily",
            report.get("daily", []),
            self.DAILY_COLUMNS,
        )
        self._write_table(
            workbook,
            "Employee Efficiency",
            report.get("employee", []),
            self.EMPLOYEE_COLUMNS,
        )
        self._write_table(
            workbook,
            "By Product",
            report.get("product", []),
            self.PRODUCT_COLUMNS,
        )
        self._write_table(
            workbook,
            "By Work Order",
            report.get("work_order", []),
            self.WORK_ORDER_COLUMNS,
        )

        temporary = target.with_name(
            f".{target.stem}.{uuid4().hex}.xlsx"
        )
        try:
            workbook.save(temporary)
            temporary.replace(target)
        finally:
            if temporary.exists():
                temporary.unlink()

        return target

    def _write_summary(
        self,
        workbook: Workbook,
        report: Mapping[str, Any],
        generated_at: datetime,
    ) -> None:
        sheet = workbook.create_sheet("Summary")
        period = dict(report.get("period", {}) or {})
        filters = dict(report.get("filters", {}) or {})
        summary = dict(report.get("summary", {}) or {})

        sheet["A1"] = "FactoryMES Manufacturing Report"
        sheet["A1"].font = self._title_font
        sheet["A1"].fill = self._title_fill
        sheet.merge_cells("A1:B1")
        sheet["A2"] = "Generated At"
        sheet["B2"] = generated_at
        sheet["B2"].number_format = "dd/mm/yyyy hh:mm:ss"
        sheet["A3"] = "Period"
        sheet["B3"] = (
            f"{period.get('start_date', '')} "
            f"to {period.get('end_date', '')}"
        )
        sheet["A4"] = "Machine Group"
        sheet["B4"] = filters.get(
            "machine_group",
            "",
        ) or "ALL"

        header_row = 6
        sheet.cell(header_row, 1, "KPI")
        sheet.cell(header_row, 2, "Value")
        self._style_header(sheet[header_row])

        for row_index, (field, label) in enumerate(
            self.SUMMARY_ROWS,
            start=header_row + 1,
        ):
            label_cell = sheet.cell(
                row_index,
                1,
                label,
            )
            value_cell = sheet.cell(
                row_index,
                2,
                self._excel_value(
                    field,
                    summary.get(field, 0),
                ),
            )
            label_cell.border = self._border
            value_cell.border = self._border
            self._apply_number_format(
                value_cell,
                field,
            )

        sheet.freeze_panes = "A7"
        self._auto_fit(sheet)

    def _write_table(
        self,
        workbook: Workbook,
        title: str,
        rows: Iterable[Any],
        columns: Sequence[tuple[str, str]],
    ) -> None:
        sheet = workbook.create_sheet(title)

        for column_index, (_, header) in enumerate(
            columns,
            start=1,
        ):
            sheet.cell(
                1,
                column_index,
                header,
            )
        self._style_header(sheet[1])

        normalized_rows = [
            dict(row)
            for row in (rows or [])
        ]

        for row_index, row in enumerate(
            normalized_rows,
            start=2,
        ):
            for column_index, (field, _) in enumerate(
                columns,
                start=1,
            ):
                cell = sheet.cell(
                    row_index,
                    column_index,
                    self._excel_value(
                        field,
                        row.get(field, ""),
                    ),
                )
                cell.border = self._border
                self._apply_number_format(cell, field)

        sheet.freeze_panes = "A2"
        last_column = get_column_letter(len(columns))
        last_row = max(1, len(normalized_rows) + 1)
        sheet.auto_filter.ref = (
            f"A1:{last_column}{last_row}"
        )
        self._auto_fit(sheet)

    def _style_header(self, cells) -> None:
        for cell in cells:
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.border = self._border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    def _apply_number_format(
        self,
        cell,
        field_name: str,
    ) -> None:
        if field_name in self.PERCENT_FIELDS:
            cell.number_format = '0.00"%"'
        elif field_name in self.INTEGER_FIELDS:
            cell.number_format = "#,##0"
        elif field_name in self.DECIMAL_FIELDS:
            cell.number_format = "#,##0.00"
        elif field_name == "production_date":
            cell.number_format = "dd/mm/yyyy"

    def _excel_value(
        self,
        field_name: str,
        value: Any,
    ) -> Any:
        if field_name in self.INTEGER_FIELDS:
            return int(round(self._number(value)))
        if (
            field_name in self.PERCENT_FIELDS
            or field_name in self.DECIMAL_FIELDS
        ):
            return round(self._number(value), 2)
        if (
            field_name == "production_date"
            and isinstance(value, (date, datetime))
        ):
            return value
        return "" if value is None else value

    @staticmethod
    def _number(value: Any) -> float:
        try:
            number = float(value or 0)
        except (TypeError, ValueError):
            return 0.0
        return number if isfinite(number) else 0.0

    @staticmethod
    def _auto_fit(sheet) -> None:
        for column_cells in sheet.columns:
            max_length = max(
                (
                    len(str(cell.value))
                    for cell in column_cells
                    if cell.value is not None
                ),
                default=0,
            )
            column_letter = get_column_letter(
                column_cells[0].column
            )
            sheet.column_dimensions[
                column_letter
            ].width = min(
                36,
                max(10, max_length + 2),
            )
