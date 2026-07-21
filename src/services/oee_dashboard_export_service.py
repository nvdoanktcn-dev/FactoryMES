from __future__ import annotations

from dataclasses import asdict, is_dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.ui.controllers.oee_dashboard_controller import OEEDashboardData


class OEEDashboardExportService:
    """
    Xuất dữ liệu OEE Dashboard sang Excel.

    Workbook gồm các sheet:
    - Summary
    - Trend
    - By Machine
    - By Employee
    - By Work Order
    - By Product
    - By Operation
    """

    SUMMARY_ROWS: Sequence[tuple[str, str]] = (
        ("execution_count", "Executions"),
        ("planned_minutes", "Planned Minutes"),
        ("runtime_minutes", "Runtime Minutes"),
        ("downtime_minutes", "Downtime Minutes"),
        ("ideal_runtime_minutes", "Ideal Runtime Minutes"),
        ("ok_quantity", "OK Quantity"),
        ("processing_ng_quantity", "Processing NG"),
        ("blank_ng_quantity", "Blank NG"),
        ("ng_quantity", "Total NG"),
        ("total_quantity", "Total Quantity"),
        ("availability", "Availability"),
        ("performance", "Performance"),
        ("quality", "Quality"),
        ("oee", "OEE"),
    )

    TREND_COLUMNS: Sequence[tuple[str, str]] = (
        ("report_date", "Date"),
        ("label", "Label"),
        ("execution_count", "Executions"),
        ("availability", "Availability"),
        ("performance", "Performance"),
        ("quality", "Quality"),
        ("oee", "OEE"),
    )

    BREAKDOWN_COLUMNS: Sequence[tuple[str, str]] = (
        ("group_key", "Group Code"),
        ("group_label", "Group"),
        ("execution_count", "Executions"),
        ("planned_minutes", "Planned Minutes"),
        ("runtime_minutes", "Runtime Minutes"),
        ("downtime_minutes", "Downtime Minutes"),
        ("ideal_runtime_minutes", "Ideal Runtime Minutes"),
        ("ok_quantity", "OK Quantity"),
        ("processing_ng_quantity", "Processing NG"),
        ("blank_ng_quantity", "Blank NG"),
        ("ng_quantity", "Total NG"),
        ("total_quantity", "Total Quantity"),
        ("availability", "Availability"),
        ("performance", "Performance"),
        ("quality", "Quality"),
        ("oee", "OEE"),
    )

    PERCENT_FIELDS = {
        "availability",
        "performance",
        "quality",
        "oee",
    }

    INTEGER_FIELDS = {
        "execution_count",
        "ok_quantity",
        "processing_ng_quantity",
        "blank_ng_quantity",
        "ng_quantity",
        "total_quantity",
    }

    DECIMAL_FIELDS = {
        "planned_minutes",
        "runtime_minutes",
        "downtime_minutes",
        "ideal_runtime_minutes",
    }

    def __init__(self) -> None:
        self._header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        self._header_font = Font(
            color="FFFFFF",
            bold=True,
        )
        self._title_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )
        self._title_font = Font(
            bold=True,
            size=14,
        )
        self._thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

    def export(
        self,
        dashboard: OEEDashboardData,
        output_path: str | Path,
        *,
        report_title: str = "OEE Dashboard Report",
        generated_at: datetime | None = None,
    ) -> Path:
        """
        Xuất dashboard ra file Excel.

        Args:
            dashboard:
                Dữ liệu đã được load từ OEEDashboardController.
            output_path:
                Đường dẫn file .xlsx.
            report_title:
                Tiêu đề báo cáo.
            generated_at:
                Thời điểm tạo báo cáo. Mặc định là datetime.now().

        Returns:
            Path của file đã tạo.
        """

        if not isinstance(dashboard, OEEDashboardData):
            raise TypeError(
                "dashboard must be an OEEDashboardData instance."
            )

        target = Path(output_path)

        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")

        target.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        generated_time = generated_at or datetime.now()

        self._write_summary_sheet(
            workbook=workbook,
            summary=dashboard.summary,
            report_title=report_title,
            generated_at=generated_time,
        )
        self._write_table_sheet(
            workbook=workbook,
            title="Trend",
            rows=dashboard.trend,
            columns=self.TREND_COLUMNS,
        )
        self._write_table_sheet(
            workbook=workbook,
            title="By Machine",
            rows=dashboard.by_machine,
            columns=self.BREAKDOWN_COLUMNS,
        )
        self._write_table_sheet(
            workbook=workbook,
            title="By Employee",
            rows=dashboard.by_employee,
            columns=self.BREAKDOWN_COLUMNS,
        )
        self._write_table_sheet(
            workbook=workbook,
            title="By Work Order",
            rows=dashboard.by_work_order,
            columns=self.BREAKDOWN_COLUMNS,
        )
        self._write_table_sheet(
            workbook=workbook,
            title="By Product",
            rows=dashboard.by_product,
            columns=self.BREAKDOWN_COLUMNS,
        )
        self._write_table_sheet(
            workbook=workbook,
            title="By Operation",
            rows=dashboard.by_operation,
            columns=self.BREAKDOWN_COLUMNS,
        )

        workbook.save(target)
        return target

    def _write_summary_sheet(
        self,
        *,
        workbook: Workbook,
        summary: Mapping[str, Any],
        report_title: str,
        generated_at: datetime,
    ) -> None:
        sheet = workbook.create_sheet(
            title="Summary"
        )

        sheet["A1"] = report_title
        sheet["A1"].font = self._title_font
        sheet["A1"].fill = self._title_fill
        sheet["A1"].alignment = Alignment(
            horizontal="left",
            vertical="center",
        )
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=2,
        )

        sheet["A2"] = "Generated At"
        sheet["B2"] = generated_at
        sheet["B2"].number_format = "dd/mm/yyyy hh:mm:ss"

        sheet["A4"] = "KPI"
        sheet["B4"] = "Value"

        for cell in sheet[4]:
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = self._thin_border

        normalized_summary = self._to_mapping(summary)

        for row_index, (
            field_name,
            label,
        ) in enumerate(
            self.SUMMARY_ROWS,
            start=5,
        ):
            label_cell = sheet.cell(
                row=row_index,
                column=1,
                value=label,
            )
            value_cell = sheet.cell(
                row=row_index,
                column=2,
                value=self._excel_value(
                    field_name,
                    normalized_summary.get(
                        field_name,
                        0,
                    ),
                ),
            )

            label_cell.border = self._thin_border
            value_cell.border = self._thin_border

            label_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
            )
            value_cell.alignment = Alignment(
                horizontal="right",
                vertical="center",
            )

            self._apply_number_format(
                cell=value_cell,
                field_name=field_name,
            )

        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = (
            f"A4:B{4 + len(self.SUMMARY_ROWS)}"
        )

        self._auto_fit_columns(sheet)

    def _write_table_sheet(
        self,
        *,
        workbook: Workbook,
        title: str,
        rows: Iterable[Any],
        columns: Sequence[tuple[str, str]],
    ) -> None:
        sheet = workbook.create_sheet(
            title=title
        )

        for column_index, (
            _,
            header,
        ) in enumerate(
            columns,
            start=1,
        ):
            cell = sheet.cell(
                row=1,
                column=column_index,
                value=header,
            )
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = self._thin_border

        normalized_rows = [
            self._to_mapping(row)
            for row in rows
        ]

        for row_index, row in enumerate(
            normalized_rows,
            start=2,
        ):
            for column_index, (
                field_name,
                _,
            ) in enumerate(
                columns,
                start=1,
            ):
                value = self._excel_value(
                    field_name,
                    row.get(
                        field_name,
                        "",
                    ),
                )

                cell = sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )
                cell.border = self._thin_border

                if field_name in {
                    "group_key",
                    "group_label",
                    "label",
                }:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                    )
                else:
                    cell.alignment = Alignment(
                        horizontal="right",
                        vertical="center",
                    )

                self._apply_number_format(
                    cell=cell,
                    field_name=field_name,
                )

        sheet.freeze_panes = "A2"

        last_column = get_column_letter(
            len(columns)
        )
        last_row = max(
            1,
            len(normalized_rows) + 1,
        )
        sheet.auto_filter.ref = (
            f"A1:{last_column}{last_row}"
        )

        self._auto_fit_columns(sheet)

    def _apply_number_format(
        self,
        *,
        cell: Any,
        field_name: str,
    ) -> None:
        if field_name in self.PERCENT_FIELDS:
            cell.number_format = '0.00"%"'
            return

        if field_name in self.INTEGER_FIELDS:
            cell.number_format = "#,##0"
            return

        if field_name in self.DECIMAL_FIELDS:
            cell.number_format = "#,##0.00"
            return

        if field_name == "report_date":
            cell.number_format = "dd/mm/yyyy"

    def _excel_value(
        self,
        field_name: str,
        value: Any,
    ) -> Any:
        if value is None:
            if field_name in self.PERCENT_FIELDS:
                return 0.0
            if field_name in self.INTEGER_FIELDS:
                return 0
            if field_name in self.DECIMAL_FIELDS:
                return 0.0
            return ""

        if field_name in self.PERCENT_FIELDS:
            return self._to_float(value)

        if field_name in self.INTEGER_FIELDS:
            return self._to_int(value)

        if field_name in self.DECIMAL_FIELDS:
            return self._to_float(value)

        if field_name == "report_date":
            return self._to_date(value)

        return value

    @staticmethod
    def _to_mapping(
        value: Any,
    ) -> dict[str, Any]:
        if value is None:
            return {}

        if isinstance(value, Mapping):
            return dict(value)

        if is_dataclass(value):
            return asdict(value)

        if hasattr(value, "_asdict"):
            return dict(value._asdict())

        if hasattr(value, "__dict__"):
            return {
                key: item
                for key, item in vars(value).items()
                if not key.startswith("_")
            }

        raise TypeError(
            (
                "Unsupported dashboard row type: "
                f"{type(value).__name__}"
            )
        )

    @staticmethod
    def _to_float(
        value: Any,
    ) -> float:
        if value in {
            None,
            "",
        }:
            return 0.0

        return round(
            float(value),
            2,
        )

    @staticmethod
    def _to_int(
        value: Any,
    ) -> int:
        if value in {
            None,
            "",
        }:
            return 0

        return int(
            round(
                float(value)
            )
        )

    @staticmethod
    def _to_date(
        value: Any,
    ) -> date | datetime | str:
        if isinstance(
            value,
            (date, datetime),
        ):
            return value

        if isinstance(value, str):
            cleaned = value.strip()

            for fmt in (
                "%Y-%m-%d",
                "%d/%m/%Y",
                "%d/%m",
            ):
                try:
                    parsed = datetime.strptime(
                        cleaned,
                        fmt,
                    )
                    return parsed.date()
                except ValueError:
                    continue

            return cleaned

        return str(value)

    @staticmethod
    def _auto_fit_columns(
        sheet: Worksheet,
        *,
        minimum_width: int = 10,
        maximum_width: int = 36,
    ) -> None:
        for column_cells in sheet.columns:
            column_letter = get_column_letter(
                column_cells[0].column
            )

            max_length = 0

            for cell in column_cells:
                value = cell.value

                if value is None:
                    continue

                if isinstance(value, datetime):
                    text = value.strftime(
                        "%d/%m/%Y %H:%M:%S"
                    )
                elif isinstance(value, date):
                    text = value.strftime(
                        "%d/%m/%Y"
                    )
                else:
                    text = str(value)

                max_length = max(
                    max_length,
                    len(text),
                )

            sheet.column_dimensions[
                column_letter
            ].width = min(
                maximum_width,
                max(
                    minimum_width,
                    max_length + 2,
                ),
            )
