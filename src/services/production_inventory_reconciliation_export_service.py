from __future__ import annotations

from datetime import date, datetime
from math import isfinite
from pathlib import Path
from typing import Any, Iterable, Mapping
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ProductionInventoryReconciliationExportService:
    """Export production-versus-finished-inventory reconciliation."""

    SUMMARY_ROWS = (
        ("work_order_count", "Công lệnh"),
        ("plan_qty", "Số lượng kế hoạch"),
        ("completed_qty", "Sản lượng OP cuối"),
        ("ng_qty", "Tổng NG"),
        ("inventory_qty", "Đã nhập kho"),
        ("pending_inventory_qty", "Còn chờ nhập kho"),
        ("over_received_qty", "Vượt nhập kho"),
        ("remaining_plan_qty", "Còn thiếu kế hoạch"),
        ("completion_percent", "Tỷ lệ hoàn thành"),
        ("inventory_percent", "Tỷ lệ nhập kho"),
        ("alert_count", "Số cảnh báo"),
    )

    RECONCILIATION_COLUMNS = (
        ("work_order_no", "Công lệnh"),
        ("product_code", "Mã sản phẩm"),
        ("plan_qty", "Kế hoạch"),
        ("completed_qty", "OP cuối"),
        ("ng_qty", "NG"),
        ("inventory_qty", "Đã nhập kho"),
        ("pending_inventory_qty", "Chờ nhập kho"),
        ("over_received_qty", "Vượt nhập"),
        ("remaining_plan_qty", "Thiếu kế hoạch"),
        ("completion_percent", "Hoàn thành"),
        ("inventory_percent", "Tỷ lệ nhập kho"),
        ("work_order_status", "Trạng thái công lệnh"),
        ("reconciliation_status", "Trạng thái đối chiếu"),
        ("last_production_date", "Ngày SX cuối"),
        ("last_inventory_date", "Ngày nhập cuối"),
    )

    DAILY_COLUMNS = (
        ("production_date", "Ngày"),
        ("completed_qty", "Sản lượng OP cuối"),
        ("inventory_qty", "Đã nhập kho"),
        ("pending_inventory_qty", "Chờ nhập kho"),
        ("over_received_qty", "Vượt nhập"),
    )

    INVENTORY_COLUMNS = (
        ("inventory_date", "Ngày nhập kho"),
        ("work_order_no", "Công lệnh"),
        ("product_code", "Mã sản phẩm"),
        ("qty", "Số lượng"),
    )

    PERCENT_FIELDS = {
        "completion_percent",
        "inventory_percent",
    }
    DATE_FIELDS = {
        "production_date",
        "inventory_date",
        "last_production_date",
        "last_inventory_date",
    }
    INTEGER_FIELDS = {
        "work_order_count",
        "plan_qty",
        "completed_qty",
        "ng_qty",
        "inventory_qty",
        "pending_inventory_qty",
        "over_received_qty",
        "remaining_plan_qty",
        "alert_count",
        "qty",
    }

    def __init__(self) -> None:
        self._header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        self._header_font = Font(
            color="FFFFFF",
            bold=True,
        )
        self._title_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )
        thin = Side(style="thin", color="D9D9D9")
        self._border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

    def export(
        self,
        report: Mapping[str, Any],
        output_path: str | Path,
    ) -> Path:
        if not isinstance(report, Mapping):
            raise TypeError("report must be a mapping.")
        target = Path(output_path)
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")
        target.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook = Workbook()
        workbook.remove(workbook.active)
        self._write_summary(workbook, report)
        self._write_table(
            workbook,
            "Reconciliation",
            report.get("rows", []),
            self.RECONCILIATION_COLUMNS,
            add_alert_formatting=True,
        )
        self._write_table(
            workbook,
            "Daily",
            report.get("daily", []),
            self.DAILY_COLUMNS,
        )
        self._write_table(
            workbook,
            "Inventory Detail",
            report.get("inventory_detail", []),
            self.INVENTORY_COLUMNS,
        )
        self._atomic_save(workbook, target)
        return target

    def _write_summary(
        self,
        workbook,
        report,
    ) -> None:
        sheet = workbook.create_sheet("Summary")
        period = dict(report.get("period", {}) or {})
        filters = dict(report.get("filters", {}) or {})
        summary = dict(report.get("summary", {}) or {})

        sheet["A1"] = (
            "FactoryMES - Đối chiếu sản xuất và nhập kho"
        )
        sheet["A1"].font = Font(
            bold=True,
            size=14,
        )
        sheet["A1"].fill = self._title_fill
        sheet.merge_cells("A1:B1")
        sheet["A2"] = "Khoảng ngày"
        sheet["B2"] = (
            f"{period.get('start_date', '')} "
            f"đến {period.get('end_date', '')}"
        )
        sheet["A3"] = "Công lệnh"
        sheet["B3"] = (
            filters.get("work_order_no")
            or "TẤT CẢ"
        )
        sheet["A4"] = "Sản phẩm"
        sheet["B4"] = (
            filters.get("product_code")
            or "TẤT CẢ"
        )
        sheet["A6"] = "Chỉ tiêu"
        sheet["B6"] = "Giá trị"
        self._style_header(sheet[6])

        for row_index, (field, label) in enumerate(
            self.SUMMARY_ROWS,
            start=7,
        ):
            sheet.cell(row_index, 1, label)
            cell = sheet.cell(
                row_index,
                2,
                self._excel_value(
                    field,
                    summary.get(field, 0),
                ),
            )
            sheet.cell(
                row_index,
                1,
            ).border = self._border
            cell.border = self._border
            self._format_cell(cell, field)

        sheet.freeze_panes = "A7"
        self._auto_fit(sheet)

    def _write_table(
        self,
        workbook,
        sheet_name,
        rows: Iterable[Any],
        columns,
        *,
        add_alert_formatting=False,
    ) -> None:
        sheet = workbook.create_sheet(sheet_name)
        for column_index, (_, label) in enumerate(
            columns,
            start=1,
        ):
            sheet.cell(1, column_index, label)
        self._style_header(sheet[1])

        normalized = [
            dict(row)
            for row in (rows or [])
        ]
        for row_index, row in enumerate(
            normalized,
            start=2,
        ):
            for column_index, (field, _) in enumerate(
                columns,
                start=1,
            ):
                cell = sheet.cell(
                    row_index,
                    column_index,
                    self._excel_value(
                        field,
                        row.get(field, ""),
                    ),
                )
                cell.border = self._border
                self._format_cell(cell, field)

        last_row = max(1, len(normalized) + 1)
        last_column = get_column_letter(
            len(columns)
        )
        sheet.auto_filter.ref = (
            f"A1:{last_column}{last_row}"
        )
        sheet.freeze_panes = "A2"
        if add_alert_formatting and normalized:
            status_column = next(
                index
                for index, (field, _) in enumerate(
                    columns,
                    start=1,
                )
                if field
                == "reconciliation_status"
            )
            status_letter = get_column_letter(
                status_column
            )
            status_range = (
                f"{status_letter}2:"
                f"{status_letter}{last_row}"
            )
            sheet.conditional_formatting.add(
                status_range,
                CellIsRule(
                    operator="equal",
                    formula=['"OVER_RECEIVED"'],
                    fill=PatternFill(
                        fill_type="solid",
                        fgColor="FFC7CE",
                    ),
                ),
            )
            sheet.conditional_formatting.add(
                status_range,
                CellIsRule(
                    operator="equal",
                    formula=['"PENDING_INVENTORY"'],
                    fill=PatternFill(
                        fill_type="solid",
                        fgColor="FFEB9C",
                    ),
                ),
            )
            sheet.conditional_formatting.add(
                status_range,
                CellIsRule(
                    operator="equal",
                    formula=['"RECONCILED"'],
                    fill=PatternFill(
                        fill_type="solid",
                        fgColor="C6EFCE",
                    ),
                ),
            )
        self._auto_fit(sheet)

    def _style_header(self, cells) -> None:
        for cell in cells:
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.border = self._border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    def _format_cell(
        self,
        cell,
        field,
    ) -> None:
        if field in self.PERCENT_FIELDS:
            cell.number_format = '0.00"%"'
        elif field in self.INTEGER_FIELDS:
            cell.number_format = "#,##0"
        elif field in self.DATE_FIELDS:
            cell.number_format = "yyyy-mm-dd"

    def _excel_value(
        self,
        field,
        value,
    ):
        if field in self.INTEGER_FIELDS:
            return int(round(self._number(value)))
        if field in self.PERCENT_FIELDS:
            return round(self._number(value), 2)
        if field in self.DATE_FIELDS:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
        return "" if value is None else value

    @staticmethod
    def _number(value) -> float:
        try:
            number = float(value or 0)
        except (TypeError, ValueError):
            return 0.0
        return number if isfinite(number) else 0.0

    @staticmethod
    def _auto_fit(sheet) -> None:
        for column_cells in sheet.columns:
            max_length = max(
                (
                    len(str(cell.value))
                    for cell in column_cells
                    if cell.value is not None
                ),
                default=0,
            )
            letter = get_column_letter(
                column_cells[0].column
            )
            sheet.column_dimensions[
                letter
            ].width = min(
                36,
                max(10, max_length + 2),
            )
        sheet.row_dimensions[1].height = 34

    @staticmethod
    def _atomic_save(
        workbook,
        target,
    ) -> None:
        temporary = target.with_name(
            f".{target.stem}.{uuid4().hex}.xlsx"
        )
        try:
            workbook.save(temporary)
            temporary.replace(target)
        finally:
            if temporary.exists():
                temporary.unlink()
