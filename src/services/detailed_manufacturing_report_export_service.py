from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, time, timedelta
from math import isfinite
from pathlib import Path
from typing import Any, Iterable, Mapping
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class DetailedManufacturingReportExportService:
    """Export production-utilization and full work-order log workbooks."""

    LOG_COLUMNS = (
        ("production_date", "Ngày"),
        ("machine_code", "Máy"),
        ("shift", "Ca"),
        ("work_order_no", "Mã CLSX"),
        ("work_order_count", "Số công lệnh"),
        ("product_code", "Mã sản phẩm"),
        ("op_no", "Công đoạn"),
        ("start_clock", "Bắt đầu"),
        ("finish_clock", "Kết thúc"),
        ("runtime_hour", "Thời gian sản xuất (giờ)"),
        ("downtime_hour", "Thời gian dừng máy (giờ)"),
        ("ok_qty", "OK"),
        ("ng_qty", "NG"),
        ("employee_code", "Tên nhân viên"),
    )

    WORK_ORDER_COLUMNS = (
        ("work_order_no", "Mã CLSX"),
        ("product_code", "Mã sản phẩm"),
        ("runtime_hour", "Tổng thời gian máy sản xuất (giờ)"),
        ("ok_qty", "Số lượng OK cuối cùng"),
        ("ng_qty", "Tổng số lượng NG (qua các công đoạn)"),
        ("total_qty", "Tổng số lượng OK + NG"),
        ("completion_date", "Ngày hoàn thành (OP cuối)"),
    )

    MACHINE_DAY_COLUMNS = (
        ("production_date", "Ngày"),
        ("machine_code", "Máy"),
        ("standard_break_min", "Nghỉ tiêu chuẩn (phút)"),
        ("runtime_min", "Tử số (phút)"),
        ("available_min", "Mẫu số (phút)"),
        ("utilization_percent", "Tỷ lệ SD"),
        ("material_wait_min", "Chờ liệu (min)"),
        ("operator_wait_min", "Chờ người (min)"),
        ("order_wait_min", "Chờ đơn (min)"),
        ("maintenance_min", "Bảo dưỡng máy móc (min)"),
        ("power_outage_min", "Mất điện (min)"),
        ("repair_min", "Sửa chữa máy (min)"),
        ("programming_min", "Sửa / Lập trình (min)"),
        ("other_min", "Khác (min)"),
        ("downtime_min", "Tổng dừng máy (phút)"),
    )

    DOWNTIME_FIELDS = (
        "material_wait_min",
        "operator_wait_min",
        "order_wait_min",
        "maintenance_min",
        "power_outage_min",
        "repair_min",
        "programming_min",
        "other_min",
    )

    DAY_BREAK_MIN = 90.0
    NIGHT_BREAK_MIN = 105.0
    DAY_PLANNED_MIN = 720.0
    NIGHT_PLANNED_MIN = 720.0

    def __init__(self) -> None:
        self._header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        self._header_font = Font(
            color="FFFFFF",
            bold=True,
        )
        self._summary_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )
        thin = Side(style="thin", color="D9D9D9")
        self._border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

    def export(
        self,
        report: Mapping[str, Any],
        output_directory: str | Path,
    ) -> tuple[Path, Path]:
        if not isinstance(report, Mapping):
            raise TypeError("report must be a mapping.")

        output_dir = Path(output_directory)
        output_dir.mkdir(parents=True, exist_ok=True)
        period_name = self._period_name(report)

        production_path = output_dir / (
            f"BaoCao_SanXuat_{period_name}.xlsx"
        )
        work_order_path = output_dir / (
            f"NhatKy_DayDuCongLenh_{period_name}.xlsx"
        )

        records = list(report.get("records", []) or [])
        self._export_production_report(
            report,
            records,
            production_path,
        )
        self._export_work_order_log(
            records,
            work_order_path,
        )
        return production_path, work_order_path

    def _export_work_order_log(
        self,
        records: list[Any],
        target: Path,
    ) -> None:
        workbook = self._new_workbook()
        grouped_records = {
            "CNC": [],
            "ROBOT": [],
        }
        for record in records:
            group = self._machine_group(
                self._text(record, "machine_code")
            )
            if group in grouped_records:
                grouped_records[group].append(record)

        for group in ("CNC", "ROBOT"):
            log_rows = [
                self._log_row(record)
                for record in self._sort_records(
                    grouped_records[group]
                )
            ]
            self._write_table(
                workbook,
                group,
                log_rows,
                self.LOG_COLUMNS,
            )

        self._write_table(
            workbook,
            "TongHop_CongLenh_CNC",
            self._work_order_rows(
                grouped_records["CNC"]
            ),
            self.WORK_ORDER_COLUMNS,
        )
        self._write_table(
            workbook,
            "TongHop_CongLenh_ROBOT",
            self._work_order_rows(
                grouped_records["ROBOT"]
            ),
            self.WORK_ORDER_COLUMNS,
        )
        self._atomic_save(workbook, target)

    def _export_production_report(
        self,
        report: Mapping[str, Any],
        records: list[Any],
        target: Path,
    ) -> None:
        workbook = self._new_workbook()
        machine_rows = self._machine_day_rows(
            report,
            records,
        )
        machines = sorted({
            row["machine_code"]
            for row in machine_rows
            if row["machine_code"]
        })

        self._write_utilization_summary(
            workbook,
            machines,
            machine_rows,
        )

        for machine_code in machines:
            self._write_table(
                workbook,
                self._safe_sheet_name(
                    machine_code
                ),
                [
                    row
                    for row in machine_rows
                    if row["machine_code"]
                    == machine_code
                ],
                self.MACHINE_DAY_COLUMNS,
            )

        self._write_monthly_utilization(
            workbook,
            machines,
            machine_rows,
        )
        self._write_downtime_by_machine(
            workbook,
            machines,
            machine_rows,
        )
        self._write_table(
            workbook,
            "Output_ByProduct",
            report.get("product", []),
            (
                ("product_code", "Mã sản phẩm"),
                ("runtime_hour", "Thời gian sản xuất (giờ)"),
                ("ok_qty", "OK cuối"),
                ("ng_qty", "NG cuối"),
                ("total_qty", "Tổng OK + NG"),
                ("yield_percent", "Tỷ lệ đạt"),
            ),
        )
        self._write_table(
            workbook,
            "Output_ByWorkOrder",
            self._work_order_rows(records),
            self.WORK_ORDER_COLUMNS,
        )
        self._atomic_save(workbook, target)

    def _machine_day_rows(
        self,
        report: Mapping[str, Any],
        records: list[Any],
    ) -> list[dict[str, Any]]:
        grouped: dict[
            tuple[str, date],
            list[Any],
        ] = defaultdict(list)
        for record in records:
            machine = self._text(
                record,
                "machine_code",
            )
            production_date = self._record_date(record)
            if machine and production_date:
                grouped[
                    (machine, production_date)
                ].append(record)

        period = dict(report.get("period", {}) or {})
        start_date = self._to_date(
            period.get("start_date")
        )
        end_date = self._to_date(
            period.get("end_date")
        )
        filters = dict(report.get("filters", {}) or {})
        selected_shift = self._normalize_shift(
            filters.get("shift")
        )

        machines = sorted({
            self._text(record, "machine_code")
            for record in records
            if self._text(record, "machine_code")
        })
        if not start_date or not end_date:
            dates = sorted({
                self._record_date(record)
                for record in records
                if self._record_date(record)
            })
        else:
            dates = [
                start_date + timedelta(days=index)
                for index in range(
                    (end_date - start_date).days + 1
                )
            ]

        rows = []
        for machine in machines:
            for production_date in dates:
                day_records = grouped.get(
                    (machine, production_date),
                    [],
                )
                downtime = {
                    field: 0.0
                    for field in self.DOWNTIME_FIELDS
                }
                for record in day_records:
                    field = self._downtime_field(
                        self._text(
                            record,
                            "downtime_reason",
                        )
                    )
                    downtime[field] += self._number(
                        getattr(
                            record,
                            "downtime_min",
                            0,
                        )
                    )

                runtime_min = sum(
                    self._number(
                        getattr(
                            record,
                            "run_time_sec",
                            0,
                        )
                    )
                    / 60
                    for record in day_records
                )
                downtime_min = sum(
                    downtime.values()
                )
                planned_min, break_min = (
                    self._planned_minutes(
                        selected_shift
                    )
                )
                available_min = max(
                    planned_min - break_min,
                    0.0,
                )
                utilization = (
                    min(
                        runtime_min
                        / available_min
                        * 100,
                        100.0,
                    )
                    if available_min > 0
                    else 0.0
                )
                row = {
                    "production_date":
                        production_date,
                    "machine_code":
                        machine,
                    "standard_break_min":
                        break_min,
                    "runtime_min":
                        runtime_min,
                    "available_min":
                        available_min,
                    "utilization_percent":
                        utilization,
                    "downtime_min":
                        downtime_min,
                }
                row.update(downtime)
                rows.append(row)
        return rows

    def _write_utilization_summary(
        self,
        workbook: Workbook,
        machines: list[str],
        rows: list[dict[str, Any]],
    ) -> None:
        sheet = workbook.create_sheet(
            "TongHop_TyLeSuDung"
        )
        metrics = (
            ("expected_min", "Thời gian mở máy dự kiến (phút)"),
            ("break_min", "Thời gian nghỉ tiêu chuẩn (phút)"),
            ("downtime_min", "Thời gian dừng máy (phút)"),
            ("runtime_min", "Thời gian mở máy thực tế (phút)"),
            ("utilization_percent", "Tỷ lệ hoạt động của thiết bị"),
        )
        sheet.cell(1, 1, "")
        for column, machine in enumerate(
            machines,
            start=2,
        ):
            sheet.cell(1, column, machine)
        total_column = len(machines) + 2
        sheet.cell(
            1,
            total_column,
            f"TỔNG {len(machines)} MÁY",
        )
        self._style_header(sheet[1])

        by_machine = defaultdict(list)
        for row in rows:
            by_machine[row["machine_code"]].append(
                row
            )

        for row_index, (field, label) in enumerate(
            metrics,
            start=2,
        ):
            sheet.cell(row_index, 1, label)
            values = []
            for column, machine in enumerate(
                machines,
                start=2,
            ):
                machine_days = by_machine[machine]
                if field == "expected_min":
                    value = sum(
                        item["available_min"]
                        + item["standard_break_min"]
                        for item in machine_days
                    )
                elif field == "break_min":
                    value = sum(
                        item["standard_break_min"]
                        for item in machine_days
                    )
                elif field == "utilization_percent":
                    denominator = sum(
                        item["available_min"]
                        for item in machine_days
                    )
                    numerator = sum(
                        item["runtime_min"]
                        for item in machine_days
                    )
                    value = (
                        min(
                            numerator
                            / denominator
                            * 100,
                            100.0,
                        )
                        if denominator
                        else 0.0
                    )
                else:
                    value = sum(
                        item[field]
                        for item in machine_days
                    )
                values.append(value)
                cell = sheet.cell(
                    row_index,
                    column,
                    value,
                )
                self._format_metric_cell(
                    cell,
                    field,
                )

            total_cell = sheet.cell(
                row_index,
                total_column,
            )
            if field == "utilization_percent":
                denominator = sum(
                    item["available_min"]
                    for item in rows
                )
                numerator = sum(
                    item["runtime_min"]
                    for item in rows
                )
                total_cell.value = (
                    min(
                        numerator
                        / denominator
                        * 100,
                        100.0,
                    )
                    if denominator
                    else 0.0
                )
            else:
                total_cell.value = sum(values)
            self._format_metric_cell(
                total_cell,
                field,
            )
        sheet.freeze_panes = "B2"
        self._auto_fit(sheet)

    def _write_monthly_utilization(
        self,
        workbook: Workbook,
        machines: list[str],
        rows: list[dict[str, Any]],
    ) -> None:
        output = []
        for machine in machines:
            machine_rows = [
                row
                for row in rows
                if row["machine_code"] == machine
            ]
            planned = sum(
                row["available_min"]
                for row in machine_rows
            )
            runtime = sum(
                row["runtime_min"]
                for row in machine_rows
            )
            downtime = sum(
                row["downtime_min"]
                for row in machine_rows
            )
            output.append({
                "machine_code": machine,
                "planned_min": planned,
                "runtime_min": runtime,
                "downtime_min": downtime,
                "utilization_percent": (
                    min(
                        runtime / planned * 100,
                        100.0,
                    )
                    if planned
                    else 0.0
                ),
                "machine_group":
                    self._machine_group(machine),
            })
        self._write_table(
            workbook,
            "Monthly_Utilization",
            output,
            (
                ("machine_code", "Máy"),
                ("machine_group", "Nhóm máy"),
                ("planned_min", "Thời gian khả dụng (phút)"),
                ("runtime_min", "Thời gian chạy (phút)"),
                ("downtime_min", "Dừng máy (phút)"),
                ("utilization_percent", "Tỷ lệ sử dụng"),
            ),
        )

    def _write_downtime_by_machine(
        self,
        workbook: Workbook,
        machines: list[str],
        rows: list[dict[str, Any]],
    ) -> None:
        output = []
        for machine in machines:
            machine_rows = [
                row
                for row in rows
                if row["machine_code"] == machine
            ]
            item = {"machine_code": machine}
            for field in self.DOWNTIME_FIELDS:
                item[field] = sum(
                    row[field]
                    for row in machine_rows
                )
            item["downtime_min"] = sum(
                item[field]
                for field in self.DOWNTIME_FIELDS
            )
            output.append(item)
        self._write_table(
            workbook,
            "Downtime_ByMachine",
            output,
            (
                ("machine_code", "Máy"),
                ("material_wait_min", "Chờ liệu"),
                ("operator_wait_min", "Chờ người"),
                ("order_wait_min", "Chờ đơn"),
                ("maintenance_min", "Bảo dưỡng"),
                ("power_outage_min", "Mất điện"),
                ("repair_min", "Sửa chữa"),
                ("programming_min", "Lập trình / Sửa"),
                ("other_min", "Khác"),
                ("downtime_min", "Tổng dừng máy"),
            ),
        )

    def _work_order_rows(
        self,
        records: Iterable[Any],
    ) -> list[dict[str, Any]]:
        grouped: dict[
            tuple[str, str],
            list[Any],
        ] = defaultdict(list)
        for record in records:
            key = (
                self._text(
                    record,
                    "work_order_no",
                ),
                self._text(
                    record,
                    "product_code",
                ),
            )
            grouped[key].append(record)

        output = []
        for (work_order, product), items in grouped.items():
            final_operation = max(
                (
                    self._operation_number(
                        self._text(item, "op_no")
                    )
                    for item in items
                ),
                default=0,
            )
            final_records = [
                item
                for item in items
                if self._operation_number(
                    self._text(item, "op_no")
                )
                == final_operation
            ]
            ok_qty = sum(
                self._integer(
                    getattr(item, "ok_qty", 0)
                )
                for item in final_records
            )
            ng_qty = sum(
                self._integer(
                    getattr(item, "ng_qty", 0)
                )
                for item in items
            )
            completion_dates = [
                self._record_date(item)
                for item in final_records
                if self._record_date(item)
            ]
            output.append({
                "work_order_no": work_order,
                "product_code": product,
                "runtime_hour": sum(
                    self._number(
                        getattr(
                            item,
                            "run_time_sec",
                            0,
                        )
                    )
                    for item in items
                )
                / 3600,
                "ok_qty": ok_qty,
                "ng_qty": ng_qty,
                "total_qty": ok_qty + ng_qty,
                "completion_date": (
                    max(completion_dates)
                    if completion_dates
                    else None
                ),
            })
        return sorted(
            output,
            key=lambda row: (
                row["work_order_no"],
                row["product_code"],
            ),
        )

    def _log_row(
        self,
        record: Any,
    ) -> dict[str, Any]:
        start = getattr(record, "start_time", None)
        finish = getattr(record, "finish_time", None)
        return {
            "production_date": self._record_date(record),
            "machine_code":
                self._text(record, "machine_code"),
            "shift": self._shift_label(
                self._text(record, "shift")
            ),
            "work_order_no":
                self._text(record, "work_order_no"),
            "work_order_count": 1,
            "product_code":
                self._text(record, "product_code"),
            "op_no": self._text(record, "op_no"),
            "start_clock": (
                start.time()
                if isinstance(start, datetime)
                else None
            ),
            "finish_clock": (
                finish.time()
                if isinstance(finish, datetime)
                else None
            ),
            "runtime_hour": self._number(
                getattr(record, "run_time_sec", 0)
            )
            / 3600,
            "downtime_hour": self._number(
                getattr(record, "downtime_min", 0)
            )
            / 60,
            "ok_qty": self._integer(
                getattr(record, "ok_qty", 0)
            ),
            "ng_qty": self._integer(
                getattr(record, "ng_qty", 0)
            ),
            "employee_code":
                self._text(record, "employee_code"),
        }

    def _write_table(
        self,
        workbook: Workbook,
        sheet_name: str,
        rows: Iterable[Any],
        columns,
    ) -> None:
        sheet = workbook.create_sheet(
            self._safe_sheet_name(sheet_name)
        )
        for column, (_, label) in enumerate(
            columns,
            start=1,
        ):
            sheet.cell(1, column, label)
        self._style_header(sheet[1])

        normalized = [
            dict(row)
            for row in (rows or [])
        ]
        for row_index, row in enumerate(
            normalized,
            start=2,
        ):
            for column_index, (field, _) in enumerate(
                columns,
                start=1,
            ):
                cell = sheet.cell(
                    row_index,
                    column_index,
                    row.get(field, ""),
                )
                cell.border = self._border
                self._format_data_cell(cell, field)

        sheet.freeze_panes = "A2"
        last_row = max(1, len(normalized) + 1)
        last_column = get_column_letter(
            len(columns)
        )
        sheet.auto_filter.ref = (
            f"A1:{last_column}{last_row}"
        )
        self._auto_fit(sheet)

    def _style_header(self, cells) -> None:
        for cell in cells:
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.border = self._border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    @staticmethod
    def _new_workbook() -> Workbook:
        workbook = Workbook()
        workbook.remove(workbook.active)
        return workbook

    def _format_data_cell(
        self,
        cell,
        field: str,
    ) -> None:
        if field in {
            "production_date",
            "completion_date",
        }:
            cell.number_format = "yyyy-mm-dd"
        elif field in {
            "start_clock",
            "finish_clock",
        }:
            cell.number_format = "h:mm"
        elif field.endswith("_percent"):
            cell.number_format = '0.00"%"'
        elif field.endswith("_hour"):
            cell.number_format = "0.00"
        elif field.endswith("_min"):
            cell.number_format = "0.00"
        elif field in {
            "ok_qty",
            "ng_qty",
            "total_qty",
            "work_order_count",
        }:
            cell.number_format = "#,##0"

    def _format_metric_cell(
        self,
        cell,
        field: str,
    ) -> None:
        cell.border = self._border
        if field == "utilization_percent":
            cell.number_format = '0.00"%"'
        else:
            cell.number_format = "#,##0.00"

    @staticmethod
    def _auto_fit(sheet) -> None:
        for column_cells in sheet.columns:
            max_length = max(
                (
                    len(str(cell.value))
                    for cell in column_cells
                    if cell.value is not None
                ),
                default=0,
            )
            letter = get_column_letter(
                column_cells[0].column
            )
            sheet.column_dimensions[
                letter
            ].width = min(
                38,
                max(10, max_length + 2),
            )
        sheet.row_dimensions[1].height = 34

    @staticmethod
    def _atomic_save(
        workbook: Workbook,
        target: Path,
    ) -> None:
        temporary = target.with_name(
            f".{target.stem}.{uuid4().hex}.xlsx"
        )
        try:
            workbook.save(temporary)
            temporary.replace(target)
        finally:
            if temporary.exists():
                temporary.unlink()

    @staticmethod
    def _sort_records(
        records: Iterable[Any],
    ) -> list[Any]:
        return sorted(
            records,
            key=lambda item: (
                getattr(item, "start_time", None)
                or datetime.min,
                str(
                    getattr(item, "machine_code", "")
                    or ""
                ),
                int(getattr(item, "id", 0) or 0),
            ),
        )

    @classmethod
    def _period_name(
        cls,
        report: Mapping[str, Any],
    ) -> str:
        period = dict(report.get("period", {}) or {})
        start_date = cls._to_date(
            period.get("start_date")
        )
        end_date = cls._to_date(
            period.get("end_date")
        )
        if start_date and end_date:
            if (
                start_date.year == end_date.year
                and start_date.month == end_date.month
            ):
                return start_date.strftime("%Y-%m")
            return (
                f"{start_date:%Y-%m-%d}"
                f"_to_{end_date:%Y-%m-%d}"
            )
        return datetime.now().strftime("%Y-%m")

    @classmethod
    def _record_date(
        cls,
        record: Any,
    ) -> date | None:
        value = getattr(record, "start_time", None)
        return cls._to_date(value)

    @staticmethod
    def _to_date(value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            try:
                return date.fromisoformat(
                    value[:10]
                )
            except ValueError:
                return None
        return None

    @staticmethod
    def _text(record: Any, field: str) -> str:
        return str(
            getattr(record, field, "")
            or ""
        ).strip().upper()

    @staticmethod
    def _number(value: Any) -> float:
        try:
            number = float(value or 0)
        except (TypeError, ValueError):
            return 0.0
        return number if isfinite(number) else 0.0

    @classmethod
    def _integer(cls, value: Any) -> int:
        return int(round(cls._number(value)))

    @staticmethod
    def _operation_number(value: str) -> int:
        digits = "".join(
            character
            for character in str(value or "")
            if character.isdigit()
        )
        return int(digits or 0)

    @staticmethod
    def _machine_group(machine_code: str) -> str:
        code = str(machine_code or "").upper()
        if code.startswith("BL"):
            return "CNC"
        if (
            code.startswith("BR")
            or code.startswith("ASK")
        ):
            return "ROBOT"
        return "OTHER"

    @staticmethod
    def _normalize_shift(value: Any) -> str:
        normalized = str(value or "").strip().upper()
        if normalized in {"DAY", "CA A", "A"}:
            return "DAY"
        if normalized in {"NIGHT", "CA B", "B"}:
            return "NIGHT"
        return ""

    @classmethod
    def _shift_label(cls, value: Any) -> str:
        normalized = cls._normalize_shift(value)
        if normalized == "DAY":
            return "Ca A"
        if normalized == "NIGHT":
            return "Ca B"
        return str(value or "").strip()

    @classmethod
    def _planned_minutes(
        cls,
        shift: str,
    ) -> tuple[float, float]:
        if shift == "DAY":
            return (
                cls.DAY_PLANNED_MIN,
                cls.DAY_BREAK_MIN,
            )
        if shift == "NIGHT":
            return (
                cls.NIGHT_PLANNED_MIN,
                cls.NIGHT_BREAK_MIN,
            )
        return (
            cls.DAY_PLANNED_MIN
            + cls.NIGHT_PLANNED_MIN,
            cls.DAY_BREAK_MIN
            + cls.NIGHT_BREAK_MIN,
        )

    @staticmethod
    def _downtime_field(reason: str) -> str:
        normalized = str(reason or "").strip().lower()
        mappings = (
            (("chờ liệu", "cho lieu", "material"), "material_wait_min"),
            (("chờ người", "cho nguoi", "operator"), "operator_wait_min"),
            (("chờ đơn", "cho don", "order"), "order_wait_min"),
            (("bảo dưỡng", "bao duong", "maintenance"), "maintenance_min"),
            (("mất điện", "mat dien", "power"), "power_outage_min"),
            (("sửa chữa", "sua chua", "repair"), "repair_min"),
            (("lập trình", "lap trinh", "program"), "programming_min"),
        )
        for keywords, field in mappings:
            if any(
                keyword in normalized
                for keyword in keywords
            ):
                return field
        return "other_min"

    @staticmethod
    def _safe_sheet_name(value: str) -> str:
        name = str(value or "Sheet").strip()
        for character in r"[]:*?/\\":
            name = name.replace(character, "_")
        return name[:31] or "Sheet"
